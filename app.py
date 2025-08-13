import streamlit as st
import pandas as pd
import io
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Table de correspondance Diametre -> Lettre pour FP2E
diametre_lettre = {
    15: ['A', 'U', 'V'],
    20: ['B'],
    25: ['C'],
    30: ['D'],
    40: ['E'],
    50: ['F'],
    60: ['G'],
    65: ['G'],
    80: ['H'],
    100: ['I'],
    125: ['J'],
    150: ['K']
}

def get_csv_delimiter(file):
    """
    Détecte automatiquement le délimiteur d'un fichier CSV.
    """
    try:
        sample = file.read(2048).decode('utf-8')
        dialect = csv.Sniffer().sniff(sample)
        file.seek(0)
        return dialect.delimiter
    except Exception:
        file.seek(0)
        return ','

# --- Fonction de vérification FP2E modifiée pour plus de précision ---
def check_fp2e_details(row):
    """
    Vérifie les détails de la norme FP2E et renvoie une chaîne détaillée
    du problème pour un coloriage précis.
    """
    anomalies = []
    
    try:
        compteur = str(row['Numéro de compteur']).strip()
        annee_fabrication_val = str(row['Année de fabrication']).strip()
        diametre_val = row['Diametre']
        
        # Vérification 1 : Format du compteur
        fp2e_regex = r'^[A-Z]\d{2}[A-Z]{2}\d{6}$'
        if not re.match(fp2e_regex, compteur):
            anomalies.append('Format de compteur non FP2E')
        else:
            annee_compteur = compteur[1:3]
            lettre_diam = compteur[4].upper()
            
            # Vérification 2 : Année de fabrication (millésime)
            if annee_fabrication_val == '' or not annee_fabrication_val.isdigit():
                anomalies.append('Année fabrication manquante ou invalide')
            else:
                annee_fabrication_padded = annee_fabrication_val.zfill(2)
                if annee_compteur != annee_fabrication_padded:
                    anomalies.append('Année millésime non conforme FP2E')
                
            # Vérification 3 : Diamètre
            fp2e_map = {'A': 15, 'U': 15, 'V': 15, 'B': 20, 'C': 25, 'D': 30, 'E': 40, 'F': 50, 'G': [60, 65], 'H': 80, 'I': 100, 'J': 125, 'K': 150}
            expected_diametres = fp2e_map.get(lettre_diam, [])
            if not isinstance(expected_diametres, list):
                expected_diametres = [expected_diametres]

            if pd.isna(diametre_val) or diametre_val not in expected_diametres:
                anomalies.append('Diamètre non conforme FP2E')

    except (TypeError, ValueError, IndexError):
        anomalies.append('Erreur de format interne')
    
    if not anomalies:
        return 'Conforme'
    else:
        return ' / '.join(anomalies)


def check_data(df):
    """
    Vérifie les données du DataFrame pour détecter les anomalies en utilisant des opérations vectorisées.
    Retourne un DataFrame avec les lignes contenant des anomalies.
    """
    df_with_anomalies = df.copy()

    df_with_anomalies['Année de fabrication'] = df_with_anomalies['Année de fabrication'].astype(str).replace('nan', '', regex=False)
    df_with_anomalies['Année de fabrication'] = df_with_anomalies['Année de fabrication'].apply(
        lambda x: str(int(float(x))) if x.replace('.', '', 1).isdigit() and x != '' else x
    )
    df_with_anomalies['Année de fabrication'] = df_with_anomalies['Année de fabrication'].str.slice(-2).str.zfill(2)
    
    # Vérification des colonnes requises
    required_columns = ['Protocole Radio', 'Marque', 'Numéro de compteur', 'Numéro de tête', 'Latitude', 'Longitude', 'Année de fabrication', 'Diametre', 'Traité', 'Mode de relève']
    if not all(col in df_with_anomalies.columns for col in required_columns):
        missing = [col for col in required_columns if col not in df_with_anomalies.columns]
        st.error(f"Colonnes requises manquantes : {', '.join(missing)}")
        st.stop()

    df_with_anomalies['Anomalie'] = ''
    df_with_anomalies['Anomalie Détaillée FP2E'] = '' # Nouvelle colonne pour les détails FP2E

    # Conversion des colonnes pour les analyses et remplacement des NaN par des chaînes vides
    df_with_anomalies['Numéro de compteur'] = df_with_anomalies['Numéro de compteur'].astype(str).replace('nan', '', regex=False)
    df_with_anomalies['Numéro de tête'] = df_with_anomalies['Numéro de tête'].astype(str).replace('nan', '', regex=False)
    df_with_anomalies['Marque'] = df_with_anomalies['Marque'].astype(str).replace('nan', '', regex=False)
    df_with_anomalies['Protocole Radio'] = df_with_anomalies['Protocole Radio'].astype(str).replace('nan', '', regex=False)
    df_with_anomalies['Traité'] = df_with_anomalies['Traité'].astype(str).replace('nan', '', regex=False)
    df_with_anomalies['Mode de relève'] = df_with_anomalies['Mode de relève'].astype(str).replace('nan', '', regex=False)
    
    # Conversion des colonnes Latitude et Longitude en numérique pour éviter le TypeError
    df_with_anomalies['Latitude'] = pd.to_numeric(df_with_anomalies['Latitude'], errors='coerce')
    df_with_anomalies['Longitude'] = pd.to_numeric(df_with_anomalies['Longitude'], errors='coerce')

    # Marqueurs pour les conditions
    is_kamstrup = df_with_anomalies['Marque'].str.upper() == 'KAMSTRUP'
    is_sappel = df_with_anomalies['Marque'].str.upper().isin(['SAPPEL (C)', 'SAPPEL (H)', 'SAPPEL(C)'])
    is_itron = df_with_anomalies['Marque'].str.upper() == 'ITRON'
    is_kaifa = df_with_anomalies['Marque'].str.upper() == 'KAIFA'
    is_mode_manuelle = df_with_anomalies['Mode de relève'].str.upper() == 'MANUELLE'
    is_mode_telerel = df_with_anomalies['Mode de relève'].str.upper() == 'TÉLERELEVE'

    annee_fabrication_num = pd.to_numeric(df_with_anomalies['Année de fabrication'], errors='coerce')
    df_with_anomalies['Diametre'] = pd.to_numeric(df_with_anomalies['Diametre'], errors='coerce')

    # ------------------------------------------------------------------
    # ANOMALIES GÉNÉRALES (valeurs manquantes et incohérences de base)
    # ------------------------------------------------------------------
    
    condition_protocole_manquant = (df_with_anomalies['Protocole Radio'].isin(['', 'nan'])) & (~is_mode_manuelle)
    df_with_anomalies.loc[condition_protocole_manquant, 'Anomalie'] += 'Protocole Radio manquant / '
    
    df_with_anomalies.loc[df_with_anomalies['Marque'].isin(['', 'nan']), 'Anomalie'] += 'Marque manquante / '
    df_with_anomalies.loc[df_with_anomalies['Numéro de compteur'].isin(['', 'nan']), 'Anomalie'] += 'Numéro de compteur manquant / '
    df_with_anomalies.loc[df_with_anomalies['Diametre'].isnull(), 'Anomalie'] += 'Diamètre manquant / '
    df_with_anomalies.loc[annee_fabrication_num.isnull(), 'Anomalie'] += 'Année de fabrication manquante / '
    
    # CORRECTION DÉFINITIVE : La condition pour Numéro de tête manquant a été mise à jour.
    # L'anomalie est levée si le Numéro de tête est manquant
    # ET la ligne ne correspond PAS aux exceptions suivantes :
    # - Kamstrup
    # - Mode Manuel
    # - Marque KAIFA en mode Télérève
    condition_tete_manquante = (df_with_anomalies['Numéro de tête'].isin(['', 'nan'])) & \
                               (~is_kamstrup) & \
                               (~is_mode_manuelle) & \
                               (~(is_kaifa & is_mode_telerel))
    df_with_anomalies.loc[condition_tete_manquante, 'Anomalie'] += 'Numéro de tête manquant / '

    df_with_anomalies.loc[df_with_anomalies['Latitude'].isnull() | df_with_anomalies['Longitude'].isnull(), 'Anomalie'] += 'Coordonnées GPS non numériques / '
    coord_invalid = ((df_with_anomalies['Latitude'] == 0) | (~df_with_anomalies['Latitude'].between(-90, 90))) | \
                    ((df_with_anomalies['Longitude'] == 0) | (~df_with_anomalies['Longitude'].between(-180, 180)))
    df_with_anomalies.loc[coord_invalid, 'Anomalie'] += 'Coordonnées GPS invalides / '

    # ------------------------------------------------------------------
    # ANOMALIES SPÉCIFIQUES AUX MARQUES
    # ------------------------------------------------------------------
    
    # KAMSTRUP
    kamstrup_valid = is_kamstrup & (~df_with_anomalies['Numéro de tête'].isin(['', 'nan']))
    df_with_anomalies.loc[is_kamstrup & (df_with_anomalies['Numéro de compteur'].str.len() != 8), 'Anomalie'] += 'KAMSTRUP: Compteur ≠ 8 caractères / '
    df_with_anomalies.loc[kamstrup_valid & (df_with_anomalies['Numéro de compteur'] != df_with_anomalies['Numéro de tête']), 'Anomalie'] += 'KAMSTRUP: Compteur ≠ Tête / '
    df_with_anomalies.loc[kamstrup_valid & (~df_with_anomalies['Numéro de compteur'].str.isdigit() | ~df_with_anomalies['Numéro de tête'].str.isdigit()), 'Anomalie'] += 'KAMSTRUP: Compteur ou Tête non numérique / '
    diametre_kamstrup_anomalie = is_kamstrup & (~df_with_anomalies['Diametre'].between(15, 80))
    df_with_anomalies.loc[diametre_kamstrup_anomalie, 'Anomalie'] += 'KAMSTRUP: Diamètre hors de la plage [15, 80] / '

    # SAPPEL
    sappel_valid = is_sappel & (~df_with_anomalies['Numéro de tête'].isin(['', 'nan']))
    df_with_anomalies.loc[sappel_valid & (df_with_anomalies['Numéro de tête'].str.len() != 16), 'Anomalie'] += 'SAPPEL: Tête ≠ 16 caractères / '
    
    # Correction pour SAPPEL C
    df_with_anomalies.loc[(is_sappel) & (df_with_anomalies['Numéro de compteur'].str.startswith('C', na=False)) & (df_with_anomalies['Marque'].str.upper() != 'SAPPEL (C)'), 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (C) / '
    
    df_with_anomalies.loc[(is_sappel) & (df_with_anomalies['Numéro de compteur'].str.startswith('H', na=False)) & (df_with_anomalies['Marque'].str.upper() != 'SAPPEL (H)'), 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (H) / '
    
    # ITRON
    itron_valid = is_itron & (~df_with_anomalies['Numéro de tête'].isin(['', 'nan']))
    df_with_anomalies.loc[itron_valid & (df_with_anomalies['Numéro de tête'].str.len() != 8), 'Anomalie'] += 'ITRON: Tête ≠ 8 caractères / '
    
    # ------------------------------------------------------------------
    # ANOMALIES MULTI-COLONNES (logique consolidée)
    # ------------------------------------------------------------------
    # Ajout de la condition pour vérifier que Protocole Radio n'est pas vide
    is_protocole_radio_filled = ~df_with_anomalies['Protocole Radio'].isin(['', 'nan'])
    
    # Protocole Radio vs Traité
    traite_lra_condition = df_with_anomalies['Traité'].str.startswith(('903', '863'), na=False)
    condition_radio_lra = traite_lra_condition & (df_with_anomalies['Protocole Radio'].str.upper() != 'LRA') & (~is_mode_manuelle) & is_protocole_radio_filled
    df_with_anomalies.loc[condition_radio_lra, 'Anomalie'] += 'Protocole ≠ LRA pour Traité 903/863 / '
    
    condition_radio_sgx = (~traite_lra_condition) & (df_with_anomalies['Protocole Radio'].str.upper() != 'SGX') & (~is_mode_manuelle) & is_protocole_radio_filled
    df_with_anomalies.loc[condition_radio_sgx, 'Anomalie'] += 'Protocole ≠ SGX pour Traité non 903/863 / '

    # ------------------------------------------------------------------
    # LOGIQUE CORRIGÉE POUR LA NORME FP2E
    # ------------------------------------------------------------------
    
    # Condition pour appliquer la vérification FP2E
    fp2e_regex = r'^[A-Z]\d{2}[A-Z]{2}\d{6}$'
    
    sappel_itron_non_manuelle = (is_sappel | is_itron) & (df_with_anomalies['Mode de relève'].str.upper() != 'MANUELLE')
    manuelle_format_ok = (df_with_anomalies['Mode de relève'].str.upper() == 'MANUELLE') & (df_with_anomalies['Numéro de compteur'].str.match(fp2e_regex, na=False))
    
    fp2e_check_condition = sappel_itron_non_manuelle | manuelle_format_ok
    
    fp2e_results = df_with_anomalies[fp2e_check_condition].apply(check_fp2e_details, axis=1)
    
    # Mise à jour des colonnes d'anomalies
    for index, result in fp2e_results.items():
        if 'Format de compteur non FP2E' in result:
            df_with_anomalies.loc[index, 'Anomalie'] += 'Format de compteur non FP2E / '
        if 'Année fabrication manquante ou invalide' in result:
            df_with_anomalies.loc[index, 'Anomalie'] += 'Année fabrication manquante ou invalide / '
        if 'Année millésime non conforme FP2E' in result:
            df_with_anomalies.loc[index, 'Anomalie'] += 'Année millésime non conforme FP2E / '
        if 'Diamètre non conforme FP2E' in result:
            df_with_anomalies.loc[index, 'Anomalie'] += 'Diamètre non conforme FP2E / '
    
    # ------------------------------------------------------------------
    # LOGIQUE DEMANDÉE : La règle "I" ou "D" pour ITRON et "C" ou "H" pour SAPPEL
    # pour les compteurs manuels FP2E
    # ------------------------------------------------------------------
    # On vérifie si le format FP2E est respecté en utilisant la regex
    is_fp2e_compliant = df_with_anomalies['Numéro de compteur'].str.match(fp2e_regex, na=False)
    
    # On applique la règle ITRON UNIQUEMENT si le mode est manuel, la marque est ITRON et le format FP2E est respecté
    itron_manuelle_fp2e_condition = is_mode_manuelle & is_itron & is_fp2e_compliant
    df_with_anomalies.loc[itron_manuelle_fp2e_condition & (~df_with_anomalies['Numéro de compteur'].str.lower().str.startswith(('i', 'd'), na=False)), 'Anomalie'] += 'ITRON manuel: doit commencer par "I" ou "D" / '
    
    # On applique la règle SAPPEL UNIQUEMENT si le mode est manuel, la marque est SAPPEL et le format FP2E est respecté
    sappel_manuelle_fp2e_condition = is_mode_manuelle & is_sappel & is_fp2e_compliant
    df_with_anomalies.loc[sappel_manuelle_fp2e_condition & (~df_with_anomalies['Numéro de compteur'].str.lower().str.startswith(('c', 'h'), na=False)), 'Anomalie'] += 'SAPPEL manuel: doit commencer par "C" ou "H" / '

    # Nettoyage de la colonne 'Anomalie'
    df_with_anomalies['Anomalie'] = df_with_anomalies['Anomalie'].str.strip().str.rstrip(' /')
    
    anomalies_df = df_with_anomalies[df_with_anomalies['Anomalie'] != ''].copy()
    anomalies_df.reset_index(inplace=True)
    anomalies_df.rename(columns={'index': 'Index original'}, inplace=True)
    
    # Comptage des anomalies pour le résumé
    anomaly_counter = anomalies_df['Anomalie'].str.split(' / ').explode().value_counts()
    
    return anomalies_df, anomaly_counter

def afficher_resume_anomalies(anomaly_counter):
    """
    Affiche un résumé des anomalies.
    """
    if not anomaly_counter.empty:
        summary_df = pd.DataFrame(anomaly_counter).reset_index()
        summary_df.columns = ["Type d'anomalie", "Nombre de cas"]
        st.subheader("Récapitulatif des anomalies")
        st.dataframe(summary_df)

# --- Interface Streamlit ---
st.title("Contrôle des données de Télérelève")
st.markdown("Veuillez téléverser votre fichier pour lancer les contrôles.")

uploaded_file = st.file_uploader("Choisissez un fichier", type=['csv', 'xlsx'])

if uploaded_file is not None:
    st.success("Fichier chargé avec succès !")

    try:
        file_extension = uploaded_file.name.split('.')[-1]
        
        # Définir le type de données pour les colonnes pour éviter la notation scientifique
        dtype_mapping = {
            'Numéro de branchement': str,
            'Abonnement': str
        }

        if file_extension == 'csv':
            delimiter = get_csv_delimiter(uploaded_file)
            df = pd.read_csv(uploaded_file, sep=delimiter, dtype=dtype_mapping)
        elif file_extension == 'xlsx':
            df = pd.read_excel(uploaded_file, dtype=dtype_mapping)
        else:
            st.error("Format de fichier non pris en charge. Veuillez utiliser un fichier .csv ou .xlsx.")
            st.stop()
    except Exception as e:
        st.error(f"Erreur de lecture du fichier : {e}")
        st.stop()

    st.subheader("Aperçu des 5 premières lignes")
    st.dataframe(df.head())

    if st.button("Lancer les contrôles"):
        st.write("Contrôles en cours...")
        anomalies_df, anomaly_counter = check_data(df)

        if not anomalies_df.empty:
            st.error("Anomalies détectées !")
            # Suppression de la colonne temporaire pour l'affichage
            anomalies_df_display = anomalies_df.drop(columns=['Anomalie Détaillée FP2E'])
            st.dataframe(anomalies_df_display)
            afficher_resume_anomalies(anomaly_counter)
            
            # Dictionnaire pour mapper les anomalies aux colonnes
            anomaly_columns_map = {
                "Protocole Radio manquant": ['Protocole Radio'],
                "Marque manquante": ['Marque'],
                "Numéro de compteur manquant": ['Numéro de compteur'],
                "Numéro de tête manquant": ['Numéro de tête'],
                "Coordonnées GPS non numériques": ['Latitude', 'Longitude'],
                "Coordonnées GPS invalides": ['Latitude', 'Longitude'],
                "Diamètre manquant": ['Diametre'],
                "Année de fabrication manquante": ['Année de fabrication'],
                
                "KAMSTRUP: Compteur ≠ 8 caractères": ['Numéro de compteur'],
                "KAMSTRUP: Compteur ≠ Tête": ['Numéro de compteur', 'Numéro de tête'],
                "KAMSTRUP: Compteur ou Tête non numérique": ['Numéro de compteur', 'Numéro de tête'],
                "KAMSTRUP: Diamètre hors de la plage [15, 80]": ['Diametre'],
                "SAPPEL: Tête ≠ 16 caractères": ['Numéro de tête'],
                "SAPPEL: Incohérence Marque/Compteur (C)": ['Numéro de compteur'],
                "SAPPEL: Incohérence Marque/Compteur (H)": ['Marque', 'Numéro de compteur'],
                "ITRON: Tête ≠ 8 caractères": ['Numéro de tête'],
                "ITRON manuel: doit commencer par \"I\" ou \"D\"": ['Numéro de compteur'],
                "SAPPEL manuel: doit commencer par \"C\" ou \"H\"": ['Numéro de compteur'],
                "Protocole ≠ LRA pour Traité 903/863": ['Protocole Radio', 'Traité'],
                "Protocole ≠ SGX pour Traité non 903/863": ['Protocole Radio', 'Traité'],
                "Format de compteur non FP2E": ['Numéro de compteur'],
                "Année millésime non conforme FP2E": ['Numéro de compteur', 'Année de fabrication'],
                "Diamètre non conforme FP2E": ['Numéro de compteur', 'Diametre'],
            }

            if file_extension == 'csv':
                csv_file = anomalies_df_display.to_csv(index=False, sep=delimiter).encode('utf-8')
                st.download_button(
                    label="Télécharger les anomalies en CSV",
                    data=csv_file,
                    file_name='anomalies_telerelève.csv',
                    mime='text/csv',
                )
            elif file_extension == 'xlsx':
                excel_buffer = io.BytesIO()
                
                wb = Workbook()
                
                default_sheet = wb.active
                wb.remove(default_sheet)
                ws_summary = wb.create_sheet(title="Récapitulatif", index=0)
                
                ws_all_anomalies = wb.create_sheet(title="Toutes_Anomalies", index=1)
                for r_df_idx, row_data in enumerate(dataframe_to_rows(anomalies_df_display, index=False, header=True)):
                    ws_all_anomalies.append(row_data)

                header_font = Font(bold=True)
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

                for cell in ws_all_anomalies[1]:
                    cell.font = header_font

                for row_num_all, df_row in enumerate(anomalies_df.iterrows()):
                    anomalies = str(df_row[1]['Anomalie']).split(' / ')
                    
                    # Logique de coloriage pour FP2E
                    fp2e_anomalies = [a for a in anomalies if 'FP2E' in a]
                    if fp2e_anomalies:
                        fp2e_details = str(df_row[1]['Anomalie Détaillée FP2E']).split(' / ')
                        
                        if 'Année fabrication différente' in fp2e_details or 'Année fabrication manquante ou invalide' in fp2e_details or 'Année millésime non conforme FP2E' in fp2e_details:
                            try:
                                col_index = list(anomalies_df_display.columns).index('Année de fabrication') + 1
                                cell = ws_all_anomalies.cell(row=row_num_all + 2, column=col_index)
                                cell.fill = red_fill
                            except ValueError:
                                pass
                        
                        if 'Diamètre non conforme FP2E' in fp2e_details:
                            try:
                                col_index = list(anomalies_df_display.columns).index('Diametre') + 1
                                cell = ws_all_anomalies.cell(row=row_num_all + 2, column=col_index)
                                cell.fill = red_fill
                            except ValueError:
                                pass
                        
                        if 'Format de compteur non FP2E' in fp2e_details or 'Erreur de format interne' in fp2e_details:
                             try:
                                col_index = list(anomalies_df_display.columns).index('Numéro de compteur') + 1
                                cell = ws_all_anomalies.cell(row=row_num_all + 2, column=col_index)
                                cell.fill = red_fill
                             except ValueError:
                                pass
                    
                    # Autres anomalies
                    for anomaly in anomalies:
                        anomaly_key = anomaly.strip()
                        if anomaly_key in anomaly_columns_map:
                            columns_to_highlight = anomaly_columns_map[anomaly_key]
                            for col_name in columns_to_highlight:
                                try:
                                    col_index = list(anomalies_df_display.columns).index(col_name) + 1
                                    cell = ws_all_anomalies.cell(row=row_num_all + 2, column=col_index)
                                    cell.fill = red_fill
                                except ValueError:
                                    pass

                for col in ws_all_anomalies.columns:
                    max_length = 0
                    column = col[0].column
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws_all_anomalies.column_dimensions[get_column_letter(column)].width = adjusted_width

                title_font = Font(bold=True, size=16)
                
                ws_summary['A1'] = "Récapitulatif des anomalies"
                ws_summary['A1'].font = title_font
                
                ws_summary.append([])
                ws_summary.append(["Type d'anomalie", "Nombre de cas"])
                ws_summary['A3'].font = header_font
                ws_summary['B3'].font = header_font
                
                created_sheet_names = set(["Récapitulatif", "Toutes_Anomalies"])

                row_num_all_anomalies = ws_summary.max_row + 1
                ws_summary.cell(row=row_num_all_anomalies, column=1, value="Toutes les anomalies").hyperlink = f"#'Toutes_Anomalies'!A1"
                ws_summary.cell(row=row_num_all_anomalies, column=1).font = Font(underline="single", color="0563C1")
                ws_summary.cell(row=row_num_all_anomalies, column=2, value=len(anomalies_df))
                ws_summary.cell(row=row_num_all_anomalies, column=2).alignment = Alignment(horizontal="right")
                
                for r_idx, (anomaly_type, count) in enumerate(anomaly_counter.items()):
                    # Logique pour raccourcir le nom de la feuille
                    sheet_name_base = anomaly_type
                    sheet_name = re.sub(r'[\\/?*\[\]:()\'"<>|]', '', sheet_name_base)
                    sheet_name = sheet_name.replace(' ', '_').replace('.', '').replace(':', '_').strip()
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:31].rstrip('_').strip()
                    
                    original_sheet_name = sheet_name
                    counter = 1
                    while sheet_name in created_sheet_names:
                        sheet_name = f"{original_sheet_name[:28]}_{counter}"
                        counter += 1
                    created_sheet_names.add(sheet_name)

                    row_num = ws_summary.max_row + 1
                    ws_summary.cell(row=row_num, column=1, value=anomaly_type)
                    ws_summary.cell(row=row_num, column=2, value=count)
                    
                    ws_anomaly_detail = wb.create_sheet(title=sheet_name)
                    
                    filtered_df = anomalies_df[anomalies_df['Anomalie'].str.contains(anomaly_type, regex=False)]
                    
                    for r_df_idx, row_data in enumerate(dataframe_to_rows(filtered_df.drop(columns=['Anomalie Détaillée FP2E']), index=False, header=True)):
                        ws_anomaly_detail.append(row_data)

                    for cell in ws_anomaly_detail[1]:
                        cell.font = header_font
                    
                    for row_num_detail, df_row in enumerate(filtered_df.iterrows()):
                        anomalies = str(df_row[1]['Anomalie']).split(' / ')

                        # Logic to handle FP2E coloring in detail sheet
                        fp2e_anomalies = [a for a in anomalies if 'FP2E' in a]
                        if fp2e_anomalies:
                            fp2e_details = str(df_row[1]['Anomalie Détaillée FP2E']).split(' / ')
                            if 'Année fabrication différente' in fp2e_details or 'Année fabrication manquante ou invalide' in fp2e_details or 'Année millésime non conforme FP2E' in fp2e_details:
                                try:
                                    col_index = list(anomalies_df_display.columns).index('Année de fabrication') + 1
                                    cell = ws_anomaly_detail.cell(row=row_num_detail + 2, column=col_index)
                                    cell.fill = red_fill
                                except ValueError:
                                    pass
                            
                            if 'Diamètre non conforme FP2E' in fp2e_details:
                                try:
                                    col_index = list(anomalies_df_display.columns).index('Diametre') + 1
                                    cell = ws_anomaly_detail.cell(row=row_num_detail + 2, column=col_index)
                                    cell.fill = red_fill
                                except ValueError:
                                    pass
                            
                            if 'Format de compteur non FP2E' in fp2e_details or 'Erreur de format interne' in fp2e_details:
                                 try:
                                    col_index = list(anomalies_df_display.columns).index('Numéro de compteur') + 1
                                    cell = ws_anomaly_detail.cell(row=row_num_detail + 2, column=col_index)
                                    cell.fill = red_fill
                                 except ValueError:
                                    pass

                        # Other anomalies
                        for anomaly in anomalies:
                            anomaly_key = anomaly.strip()
                            if anomaly_key in anomaly_columns_map:
                                columns_to_highlight = anomaly_columns_map[anomaly_key]
                                for col_name in columns_to_highlight:
                                    try:
                                        col_index = list(anomalies_df_display.columns).index(col_name) + 1
                                        cell = ws_anomaly_detail.cell(row=row_num_detail + 2, column=col_index)
                                        cell.fill = red_fill
                                    except ValueError:
                                        pass

                    for col in ws_anomaly_detail.columns:
                        max_length = 0
                        column = col[0].column
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws_anomaly_detail.column_dimensions[get_column_letter(column)].width = adjusted_width

                    ws_summary.cell(row=row_num, column=1).hyperlink = f"#'{sheet_name}'!A1"
                    ws_summary.cell(row=row_num, column=1).font = Font(underline="single", color="0563C1")
                    
                for col in ws_summary.columns:
                    max_length = 0
                    column = col[0].column
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws_summary.column_dimensions[get_column_letter(column)].width = adjusted_width
                    
                excel_buffer_styled = io.BytesIO()
                wb.save(excel_buffer_styled)
                excel_buffer_styled.seek(0)

                st.download_button(
                    label="Télécharger les anomalies en Excel",
                    data=excel_buffer_styled,
                    file_name='anomalies_telerelève.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
        else:
            st.success("Aucune anomalie détectée ! Les données sont conformes.")
